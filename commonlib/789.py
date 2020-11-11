import openpyxl
import os
import requests
import urllib
import urllib.error
from openpyxl import load_workbook
requestdata1_tmp='<?xml version="1.0" encoding="GBK"?>\
<operation_in>\
  <verify_code>304196201703300000000089</verify_code>\
  <sysfunc_id>10100003</sysfunc_id>\
  <process_code>OPC_DeliveryPrepareGoods</process_code>\
  <LOG_LEVEL>5</LOG_LEVEL>\
  <accept_info>\
    <accept_city>591</accept_city>\
    <accept_org_id>1110314</accept_org_id>\
    <accept_province>5910</accept_province>\
  </accept_info>\
  <request_source>201028</request_source>\
  <log_info>\
    <tx_id>301158341570092871470113</tx_id>\
    <msg_id>1</msg_id>\
  </log_info>\
  <request_time>20191003165431</request_time>\
  <content>\
    <delivery_id>100000006635</delivery_id>\
    <logistics_company_id>10000409</logistics_company_id>\
    <logistics_company_name>中东快递</logistics_company_name>\
    <supplier_addr>测试-国货路营业厅24小时自助营业厅</supplier_addr>\
    <modify_id>1021439</modify_id>\
  </content>\
</operation_in>\
'
#url="http://10.45.138.39:8380/opc-core-microservice/order/orderCancel"
url="http://210.45.138.39:8380/opc-core-microservice/delivery/prepareGoods"
url2="http://10.45.138.39:8380/opc-core-microservice/delivery/setSupplier"

requestdata1=requestdata1_tmp.encode("utf-8")

def request_error(url, request_data):
    response_data = ""
    error_code = ""
    error_text = ""
    try:
        # urllib.request.urlopen("https://www.cnblogs.com/mcq1999/p/python_Crawler_1.html")
        urllib.request.urlopen(url, request_data)  # request_data入参必须为utf-8类型的数据
        response_data = "OK"
    except urllib.error.URLError as e:  # 捕捉错误时有可能只有出错原因，没有出错代码
        if hasattr(e, "code"):
            error_code = str(e.code)
            # print(error_code)
        if hasattr(e, "reason"):
            error_text = str(e.reason)
            #print(error_text)
        if error_code != "":
            pass
            #response_data = "错误代码:" + error_code + "  错误原因:" + error_text
        else:
            pass
            response_data ="错误原因:"+error_text
    else:
         response_data = "没有获取到正确的请求url地址，请核对url地址在表格当中的行号列号入参"
    print(response_data)

request_error(url,requestdata1)










# try:
#     urllib.request.urlopen(url,requestdata1.encode("utf-8"))
# except urllib.error.URLError as e:
#     if hasattr(e,"code"):
#         #print(e.code)
#         print("代码：",str(e.code))
#     if hasattr(e,"reason"):
#         print("原因：",e.reason)
#requests.post(url, data="123")

# # BASE_DIR = os.path.abspath(os.path.dirname(__file__))
# # file_dir = os.path.join(BASE_DIR, '77.py')
#
# excel_path="D:\PythonProject\data_excel\五中心测试101.xlsx"
# book_excel=load_workbook(excel_path)
# sheet_excel=book_excel.worksheets[1]
# sheet_excel.cell(3,1).value="敌人1001"
# sheet_excel.cell(3,2).value="敌人2002"
# sheet_excel.cell(3,3).value="敌人3003"
# book_excel.save(excel_path)
