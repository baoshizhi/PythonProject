#encoding=utf-8
import openpyxl
import re
from openpyxl import load_workbook
from commonlib.get_textvalue import GetTextValue
class ReadFromExcel:
    def __init__(self):
        print("__init__")
    def __del__(self):
        print("__def__")
    '''
    excel_path:接口测试excel文件名
    sheet_index:excel表格中的sheet表序号(计数从1开始)
    request_host_col：excel表中request请求目标主机地址所在的列数(列计数从1开始)
    request_type_col：excel表中request请求接口类型(post/get)所在的列数(列计数从1开始)
    request_text_col：excel表中request请求数据所在的列数(列计数从1开始)
    expect_col：excel表中request请求后返回的数据预期值所在的列数(列计数从1开始)
    start_row：excel表中剔除字段值后的有效数据所在行数(行计数从1开始)
    row_i:要获取的目标数据所在行的行号
    函数指定的待返回数据resp_type分为(host/send_type/send_data/exp_data)
    '''
    #打开接口测试excel表中的sheet子表名，sheet_index:excel表格中的sheet表序号(计数从1开始)
    def open_excel(self,excel_path,sheet_index):
        # 打开excel表
        self.excel_path=excel_path
        self.book_excel=load_workbook(self.excel_path)
        # 定位excel表中的某个sheet字表
        self.sheet_index=sheet_index-1
        self.sheet_excel=self.book_excel.worksheets[self.sheet_index]

    #获取请求相关的数据如主机地址，请求类型，请求数据文本，预期数据，
    def get_request_data(self,request_host_col,request_type_col,request_text_col,expect_text_col,row_i,resp_type):
            #获取request请求目标主机地址所在excel单元格文本信息
            text_temp_1=self.sheet_excel.cell(row_i,request_host_col).value
            #print("text_temp_1:---",text_temp_1)
            #匹配http开头直到文本结束的文本(list变量类型)，本身不包含http字符，通过(变量名[下标值])取值
            text_temp_2= re.findall("http(.+?)$",text_temp_1)
            #print("text_tmp_2:----",text_temp_2)
            #往缺少http字符开头的字符串补添加http字符
            text_temp="http"+text_temp_2[0]  #list变量类型 通过下标值取值【数字】取值
            #print("text_temp:----",text_temp)
            self.request_host=text_temp
            print("主机地址:\n",self.request_host)

            #获取request请求类型，一般为post或者get
            self.request_type=self.sheet_excel.cell(row_i,request_type_col).value
            #print("request请求类型",self.request_type)

            #获取请求体数据文本
            self.request_text=self.sheet_excel.cell(row_i,request_text_col).value
            #print("request请求体文本:\n",self.request_text)

            #获取预期数据文本
            self.expect_text_tmp=self.sheet_excel.cell(row_i,expect_text_col).value
            #去除预期数据中的空白字符得到纯净的文本内容以方便用来与返回回来的标准文本进行比较
            #通过sub函数将文本中所有的空白字符全部替换为空字符，"\s"表示空白字符
            self.expect_text=re.sub("\s","",self.expect_text_tmp)
            #print("预期文本:\n",self.expect_text)
            #返回指定的数据
            if  resp_type=="host":
                #print("主机地址:\n", self.request_host)
                return self.request_host
            elif resp_type=="send_type":
                #print("request请求类型", self.request_type)
                return self.request_type
            elif resp_type=="send_data":
                #print("request请求文本:\n", self.request_text)
                return  self.request_text
            elif resp_type=="exp_data":
                #print("去除空白内容格式化后的初始预期文本:\n", self.expect_text)
                return self.expect_text
    #往之前打开过的源表下指定的序号sheet_index子sheet表(计数从1开始)的单元格写入text文本数据
    #单元格行号row_i与列号col_j计数均从1开始
    def write_excel(self,sheet_index,row_i,col_j,text):
        self.sheet_excel.cell(row_i, col_j).value=text
        self.book_excel.save(self.excel_path)
        # # 预定一个字体相关的格式对象style
        # style=xlwt.XFStyle()
        # # 字体参数相关配置
        # font=xlwt.Font()
        # font.name="Microsoft YaHei UI"
        # font.bold=True  # 是否加粗，True为加粗  False为不加粗
        # font.height=250  # 如果是200对应的是10号字体
        # style.font=font
        # # 设置单元格对齐方式
        # alignment = xlwt.Alignment()
        # # 0x01(左端对齐)、0x02(水平方向上居中对齐)、0x03(右端对齐)
        # alignment.horz = 0x01 #水平方向左端对齐
        # # 0x00(上端对齐)、 0x01(垂直方向上居中对齐)、0x02(底端对齐)
        # alignment.vert = 0x00 #垂直方向上端对齐
        # alignment.wrap = 1 #自动换行
        # style.alignment=alignment
        # # 边框
        # borders=xlwt.Borders()
        # borders.top=xlwt.Borders.THIN
        # borders.bottom=xlwt.Borders.THIN
        # borders.left=xlwt.Borders.THIN
        # borders.right=xlwt.Borders.THIN
        # style.borders=borders
       #  # 写入数据
       #  '''
       #  如果写入数据之前原表格当中存在日期类型数据，则调用save函数保存时会报类型错误，无法保存。
       #  TypeError: descriptor 'decode' requires a 'bytes' object but received a 'NoneType'
       #  解决办法：
       #  手动打开excel表格将所有sheet子表下的日期类型数据所在列的单元格的格式重新设为日期格式或设为文本格式。
       # '''
       #  # new_sheet.write(row_i-1,col_j-1,text,style)
       #  # col_size = new_sheet.col(col_j-1)
       #  # col_size.width=320*20  #固定写入数据的所在单元格的行高20,宽度320
       #  #另存为excel文件，并将文件命名如果与原来文件重名则修改原文件


if __name__ == '__main__':
    # 实例化一个读取excel文本类
    readexcel = ReadFromExcel()
    # #打开表
    #readexcel.open_excel("D:\PythonProject\data_excel\测试用例.xls",1)
    readexcel.open_excel("D:\PythonProject\data_excel\五中心测试101.xlsx", 2)
    # 获取请求的相关数据,主机host,请求类型send_type,请求数据send_data,预期文本exp_data
    host=readexcel.get_request_data(9, 10, 11, 12, 3,"host")
    send_type=readexcel.get_request_data(9, 10, 11, 12, 3, "send_type")
    send_data=readexcel.get_request_data(9, 10, 11, 12, 3, "send_data")
    exp_data=readexcel.get_request_data(9, 10, 11, 12, 3, "exp_data")
    #readexcel.write_excel("D:\PythonProject\data_excel\测试用例.xls",1,3,15,"123")
    readexcel.write_excel(2,3,15,"7777777777")
    print("host:",host)
    print("send_type:",send_type)
    print("send_data:\n",send_data)
    print("exp_data:\n",exp_data)















