#encoding=utf-8
import openpyxl
from openpyxl.styles import  Font,Alignment,Border,Side,Color
import re

'''
def get_request_data(self, request_host_col, request_type_col, request_text_col, expect_text_col, 
                           row_i, resp_type)
    request_host_col：excel表中request请求目标主机地址所在的列数(列计数从1开始)
    request_type_col：excel表中request请求接口类型(post/get)所在的列数(列计数从1开始)
    request_text_col：excel表中request请求数据所在的列数(列计数从1开始)
    expect_col：excel表中request请求后返回的数据预期值所在的列数(列计数从1开始)
    row_i:要获取的目标数据所在行的行号
    函数指定的待返回数据resp_type分为(host/send_type/send_data/exp_data)      
    
def request(self, request_type,url, para,resp_source="no")     
     1)request_type:根据传入的变量请求类型【post/get】自动发起相应请求,
                    resp_source值默认为no,为yes时返回未被去除空白字符的初始数据
     2)url:向服务器发起请求的地址链接
     3)para:向服务器发起请求的请求体数据
'''

class ReadWriteExcel:
    def __init__(self):
        print("__init__")
    def __del__(self):
        print("__def__")
    #打开接口测试excel表中的某个sheet子表名，并返回sheet子表的总个数
    #excel_path: 接口测试excel表格绝对路径文件名
    def open_excel(self,excel_path,sheet_num=0):
        # 打开excel表
        self.excel_path = excel_path
        self.book_excel = openpyxl.load_workbook(self.excel_path)
        #print("book_excel:",self.book_excel)
        self.sheetnames=self.book_excel.sheetnames #返回的类型为list变量
        #print("sheet表总个数:",len(self.sheetnames))
        if sheet_num==0:
            return len(self.sheetnames)
        else:
            return  self.book_excel

    #sheet_index: excel表格中的sheet表序号(计数从1开始)，并返回sheet表的总行数
    def open_sheet(self,sheet_index):
        # 定位excel表中的某个sheet字表
        self.sheet_index = sheet_index - 1
        print(self.book_excel)
        self.sheet_excel = self.book_excel.worksheets[self.sheet_index]
        #获取该表总行数
        self.lines_total=self.sheet_excel.max_row
        return self.lines_total

    #获取请求相关的数据如主机地址，请求类型，请求数据文本，预期数据，
    def get_request_data(self,request_host_col,request_type_col,request_text_col,expect_text_col,row_i,resp_type):
            #获取request请求目标主机地址所在excel单元格文本信息
            text_temp_1=self.sheet_excel.cell(row_i,request_host_col).value
            if text_temp_1==None or text_temp_1=="":
                print("未获取到主机地址，请检查表格数据")
                return
            #print("text_temp_1:---",text_temp_1)
            #匹配http开头直到文本结束的文本(list变量类型)，本身不包含http字符，通过(变量名[下标值])取值
            text_temp_2= re.findall("http(.+?)$",text_temp_1)
            #print("text_tmp_2:----",text_temp_2)
            if text_temp_2==[]:
                text_temp_2=[""]
                #print("未获取到http地址,请检查地址字段所在的行号列号入参值")
            #往缺少http字符开头的字符串补添加http字符
            text_temp="http"+text_temp_2[0]  #list变量类型 通过下标值取值【数字】取值
            #去除主机地址中为空的字符(尾部空字符)
            text_temp=re.sub("\s","",text_temp)
            self.request_host=text_temp
            #print("主机地址:\n",self.request_host)

            #获取request请求类型，一般为post或者get
            self.request_type=self.sheet_excel.cell(row_i,request_type_col).value
            type=["Post","post","Get","get"]
            if self.request_type==None or self.request_type not in type:
                #print("未获取到支持的请求类型post/get，请检查表格数据")
                return
            #print("request请求类型",self.request_type)

            #获取请求体数据文本
            self.request_text=self.sheet_excel.cell(row_i,request_text_col).value
            if self.request_text == None or self.request_text == "":
                #print("未获取到请求体数据类型，请检查表格数据")
                return
            #print("request请求体文本:\n",self.request_text)

            #获取预期数据文本
            self.expect_text_tmp=self.sheet_excel.cell(row_i,expect_text_col).value
            if self.expect_text_tmp == None or self.expect_text_tmp == "":
                #print("未获取到预期检查点数据文本，请检查表格数据")
                return
            #print(self.expect_text_tmp)
            #去除预期数据中的空白字符得到纯净的文本内容以方便用来与返回回来的标准文本进行比较
            #通过sub函数将文本中所有的空白字符全部替换为空字符，"\s"表示空白字符
            self.expect_text=re.sub("\s","",self.expect_text_tmp)
            #print("预期文本:\n",self.expect_text)
            #返回指定的数据
            if  resp_type=="host":
                #print("主机地址:\n", self.request_host)
                return self.request_host
            elif resp_type=="send_type":
                #print("request请求类型", self.request_type)
                return self.request_type
            elif resp_type=="send_data":
                #print("request请求文本:\n", self.request_text)
                return  self.request_text
            elif resp_type=="exp_data":
                #print("去除空白内容格式化后的初始预期文本:\n", self.expect_text)
                return self.expect_text

    #往之前打开过的源表下指定的序号sheet_index子sheet表(计数从1开始)的单元格写入text文本数据
    #单元格行号row_i与列号col_j计数均从1开始
    def write_excel(self,sheet_index,row_i,col_j,text):
        # 字体参数相关配置
        font=Font()
        font.name=u"宋体"
        font.bold=True  # 是否加粗，True为加粗  False为不加粗
        font.size=12  # 字体大小
        font.italic=False #是否斜体，True斜体
        if text=="未通过" :  #测试未通过 红色字体显示
            font.color="FF0000" #黑色000000  红色FF0000  蓝色00008B
        elif text[0:18]=="python内置requests函数" or text[0:14]=="未获取到表格中正确的相关数据":
            font.color = "00008B"  # 黑色000000  红色FF0000  蓝色00008B
        self.sheet_excel.cell(row_i, col_j).font=font

        # 设置单元格对齐方式
        #horizontal代表水平方向，左对齐left，右对齐right,居中center，分散对齐distributed，跨列居中centerContinuous，两端对齐justify，填充fill，常规general
        #vertical代表垂直方向，靠上top，居中center，靠下bottom，两端对齐justify，分散对齐distributed
        #wrap_text代表自动换行，True和False
        align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        self.sheet_excel.cell(row_i, col_j).alignment = align

        # 边框
        borders=Border()
        borders.top=Side(border_style='thin',color='000000')
        borders.bottom = Side(border_style='thin', color='000000')
        borders.left = Side(border_style='thin', color='000000')
        borders.right = Side(border_style='thin', color='000000')
        self.sheet_excel.cell(row_i, col_j).border = borders

        self.sheet_excel.row_dimensions[row_i].height=20 #行高20
        #self.sheet_excel.column_dimensions[col_j].width=320.0 #宽度320
        # 写入数据
        self.sheet_excel.cell(row_i, col_j).value = text  # 写入传入的数据
        #self.book_excel.save(self.excel_path) #写完所有的sheet表的数据后再统一一次保存excel表(边写边保存效率低)











